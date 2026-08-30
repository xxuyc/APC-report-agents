"""Deterministically extract project-identification tokens from input files."""

from __future__ import annotations

import re
from pathlib import Path


def collect_input_tokens(paths, max_cells=1200):
    tokens = set()
    for raw in paths:
        path = Path(raw)
        tokens.add(path.stem.strip().lower())
        tokens.update(x.lower() for x in re.split(r"[-_（）()\s]+", path.stem) if len(x) > 1)
        if path.suffix.lower() != ".xlsx" or not path.is_file():
            continue
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(path, read_only=True, data_only=True)
            count = 0
            for sheet in workbook.worksheets:
                tokens.add(sheet.title.strip().lower())
                for row in sheet.iter_rows(max_row=80, max_col=20, values_only=True):
                    for value in row:
                        if isinstance(value, str) and 2 <= len(value.strip()) <= 120:
                            tokens.add(value.strip().lower())
                        count += 1
                        if count >= max_cells: break
                    if count >= max_cells: break
                if count >= max_cells: break
            workbook.close()
        except Exception:
            # Filename matching remains available when optional workbook parsing fails.
            continue
    return tokens
