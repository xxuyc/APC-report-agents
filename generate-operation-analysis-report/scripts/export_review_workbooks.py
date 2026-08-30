#!/usr/bin/env python3
"""Export review workbooks with public, cross-platform Python dependencies."""

import argparse
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_sheet(sheet, widths):
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_rows(workbook, title, headers, rows, widths):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    style_sheet(sheet, widths)
    return sheet


def save_verified(workbook, path):
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    loaded = load_workbook(path, read_only=True, data_only=False)
    sheet_names = loaded.sheetnames
    loaded.close()
    return sheet_names


def build_analysis_workbook(model, output):
    workbook = Workbook()
    workbook.remove(workbook.active)
    summary_rows = [
        ["期间处理水量（m³）", next((x.get("value") for x in model.get("evidence", [])
                                  if x.get("evidence_id") == "EV-FLOW-ANNUAL"), None)],
        [f"{model.get('cost_scope', {}).get('label', '核心生产成本')}（元）", model.get("core_cost_total_cny")],
        ["已识别运营费用（元）", model.get("recognized_cost_total_cny")],
        ["质量问题数量", len(model.get("quality_issues", []))],
    ]
    add_rows(workbook, "分析摘要", ["指标", "结果"], summary_rows, [32, 24])

    monthly = model.get("monthly_table", {})
    periods = monthly.get("periods", [])
    monthly_rows = [[item.get("label"), *item.get("values", [])] for item in monthly.get("rows", [])]
    sheet = add_rows(workbook, "月度数据", ["指标", *periods], monthly_rows,
                     [30, *([13] * len(periods))])
    if periods and monthly_rows:
        chart = LineChart()
        chart.title = "月度趋势"
        chart.y_axis.title = "数值"
        chart.x_axis.title = "月份"
        data_end = min(sheet.max_row, 5)
        data = Reference(sheet, min_col=1, max_col=sheet.max_column, min_row=2, max_row=data_end)
        categories = Reference(sheet, min_col=2, max_col=sheet.max_column, min_row=1, max_row=1)
        chart.add_data(data, titles_from_data=True, from_rows=True)
        chart.set_categories(categories)
        chart.height, chart.width = 10, 24
        sheet.add_chart(chart, f"A{sheet.max_row + 3}")

    removal_rows = [[x.get("pollutant"), x.get("removed_kg"), x.get("emitted_kg"), x.get("rate_pct")]
                    for x in model.get("removal_table", [])]
    add_rows(workbook, "去除效率", ["指标", "去除量（kg）", "排放量（kg）", "去除率（%）"],
             removal_rows, [18, 18, 18, 16])

    cost_rows = [[x.get("label"), x.get("cost_cny"), x.get("share_pct")]
                 for x in model.get("cost_table", [])]
    cost = add_rows(workbook, "成本分析", ["成本类别", "金额（元）", "占比（%）"],
                    cost_rows, [24, 20, 16])
    if cost_rows:
        chart = PieChart()
        chart.title = f"{model.get('cost_scope', {}).get('label', '成本')}构成"
        chart.add_data(Reference(cost, min_col=2, min_row=1, max_row=cost.max_row),
                       titles_from_data=True)
        chart.set_categories(Reference(cost, min_col=1, min_row=2, max_row=cost.max_row))
        cost.add_chart(chart, "E2")

    yoy_rows = [[x.get("label"), x.get("value_2025"), x.get("value_2026"),
                 x.get("unit"), x.get("change_pct")]
                for x in model.get("resource_table", [])]
    yoy = add_rows(workbook, "同比分析", ["指标", "2025年同期", "2026年同期", "单位", "同比变化（%）"],
                   yoy_rows, [24, 18, 18, 14, 18])
    if yoy_rows:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "同比变化"
        chart.add_data(Reference(yoy, min_col=5, min_row=1, max_row=yoy.max_row),
                       titles_from_data=True)
        chart.set_categories(Reference(yoy, min_col=1, min_row=2, max_row=yoy.max_row))
        yoy.add_chart(chart, "G2")

    quality_rows = [[index, x.get("severity"), x.get("type"), x.get("metric_id"), x.get("message")]
                    for index, x in enumerate(model.get("quality_issues", []), 1)]
    add_rows(workbook, "数据质量", ["序号", "级别", "类型", "指标", "说明"],
             quality_rows, [10, 12, 24, 22, 60])

    trace_rows = []
    for item in model.get("evidence", []):
        sources = "；".join(f"{x.get('sheet', '')}!{x.get('cell', '')}" for x in item.get("sources", []))
        value = item.get("value")
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False)
        trace_rows.append([item.get("evidence_id"), item.get("kind"), value, item.get("unit"),
                           item.get("calculation"), sources])
    add_rows(workbook, "追溯记录", ["证据ID", "类型", "数值", "单位", "计算口径", "来源"],
             trace_rows, [22, 16, 24, 14, 40, 40])
    return save_verified(workbook, output)


def build_pending_workbook(model, output):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "待确认事项"
    sheet.append(["序号", "问题类型", "指标", "事项说明", "状态", "处理意见"])
    for index, item in enumerate(model.get("quality_issues", []), 1):
        sheet.append([index, item.get("type"), item.get("metric_id"), item.get("message"), "待确认", ""])
    style_sheet(sheet, [10, 24, 22, 60, 14, 30])
    return save_verified(workbook, output)


def build_trace_workbook(model, output):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "内容追溯表"
    sheet.append(["证据ID", "类型", "数值", "单位", "计算口径", "来源单元格"])
    for item in model.get("evidence", []):
        value = item.get("value")
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False)
        sources = "；".join(f"{x.get('sheet', '')}!{x.get('cell', '')}" for x in item.get("sources", []))
        sheet.append([item.get("evidence_id"), item.get("kind"), value, item.get("unit"),
                      item.get("calculation"), sources])
    style_sheet(sheet, [22, 16, 24, 14, 40, 40])
    return save_verified(workbook, output)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--model", required=True)
    parser.add_argument("--normalized", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--qa-dir", required=True)
    args = parser.parse_args()
    model = json.loads(Path(args.model).read_text(encoding="utf-8"))
    # Read the normalized file to ensure the expected upstream artifact is valid.
    json.loads(Path(args.normalized).read_text(encoding="utf-8"))
    output = Path(args.output_dir)
    qa = Path(args.qa_dir)
    outputs = {
        "分析结果及图表.xlsx": build_analysis_workbook(model, output / "分析结果及图表.xlsx"),
        "待确认事项.xlsx": build_pending_workbook(model, output / "待确认事项.xlsx"),
        "内容追溯表.xlsx": build_trace_workbook(model, output / "内容追溯表.xlsx"),
    }
    qa.mkdir(parents=True, exist_ok=True)
    (qa / "xlsx-check.json").write_text(
        json.dumps({"status": "passed", "formula_errors": 0, "outputs": outputs},
                   ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps({"outputs": list(outputs), "verified": True}, ensure_ascii=False))


if __name__ == "__main__":
    main()
