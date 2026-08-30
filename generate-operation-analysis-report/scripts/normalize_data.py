#!/usr/bin/env python3
import argparse
import calendar
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from common import month_label, read_json, write_json


def numeric(value):
    if isinstance(value, bool) or value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", ""))
    except ValueError:
        return None


def normalize_two_plant(args, adapter, profile, formulas, values):
    sheet_name = adapter["sheet"]
    ws_f, ws_v = formulas[sheet_name], values[sheet_name]
    year = int(profile["analysis_year"])
    periods = [f"{year:04d}-{int(block['month']):02d}" for block in adapter["period_blocks"]]
    cols = {key: column_index_from_string(value) for key, value in adapter["columns"].items() if key != "note"}
    note_col = column_index_from_string(adapter["columns"]["note"])

    raw = {"south": [], "east": []}
    for block in adapter["period_blocks"]:
        for plant_key, row_key in (("south", "south_row"), ("east", "east_row")):
            row = int(block[row_key])
            item = {key: numeric(ws_v.cell(row, col).value) for key, col in cols.items() if key != "plant"}
            item.update({"row": row, "month": int(block["month"]), "note": ws_v.cell(row, note_col).value})
            raw[plant_key].append(item)

    metric_defs = {
        "total_flow_m3": ("m3", "sum"), "total_flow_10k_m3": ("10k_m3", "sum"),
        "south_flow_m3": ("m3", "sum"), "east_flow_m3": ("m3", "sum"),
        "south_flow_10k_m3": ("10k_m3", "sum"), "east_flow_10k_m3": ("10k_m3", "sum"),
        "daily_flow_10k_m3d": ("10k_m3_d", "mean"), "south_load_pct": ("pct", "mean"), "east_load_pct": ("pct", "mean"),
        "electricity_kwh": ("kWh", "sum"), "south_electricity_kwh": ("kWh", "sum"), "east_electricity_kwh": ("kWh", "sum"),
        "energy_intensity_kwh_m3": ("kWh_m3", "mean"),
        "south_energy_intensity_kwh_m3": ("kWh_m3", "mean"), "east_energy_intensity_kwh_m3": ("kWh_m3", "mean"),
        "carbon_source_usage_kg": ("kg", "sum"), "carbon_intensity_kg_m3": ("kg_m3", "mean"),
        "flocculant_usage_kg": ("kg", "sum"), "hypochlorite_usage_kg": ("kg", "sum"),
        "hypochlorite_intensity_kg_m3": ("kg_m3", "mean"), "tap_water_m3": ("m3", "sum"),
        "sludge_t": ("t", "sum"), "pac_raw_kg": ("kg", "sum")
    }
    series_values = {key: [] for key in metric_defs}
    sources = {key: [] for key in metric_defs}
    for idx, block in enumerate(adapter["period_blocks"]):
        s, e = raw["south"][idx], raw["east"][idx]
        days = calendar.monthrange(year, int(block["month"]))[1]
        flow = (s["total_flow_m3"] or 0) + (e["total_flow_m3"] or 0)
        power = (s["electricity_kwh"] or 0) + (e["electricity_kwh"] or 0)
        carbon = (s["carbon_source_usage_kg"] or 0) + (e["carbon_source_usage_kg"] or 0)
        hypo = (s["hypochlorite_usage_kg"] or 0) + (e["hypochlorite_usage_kg"] or 0)
        values_map = {
            "total_flow_m3": flow, "total_flow_10k_m3": flow / 10000,
            "south_flow_m3": s["total_flow_m3"], "east_flow_m3": e["total_flow_m3"],
            "south_flow_10k_m3": s["total_flow_m3"] / 10000, "east_flow_10k_m3": e["total_flow_m3"] / 10000,
            "daily_flow_10k_m3d": flow / days / 10000,
            "south_load_pct": s["load_rate"] * 100, "east_load_pct": e["load_rate"] * 100,
            "electricity_kwh": power, "south_electricity_kwh": s["electricity_kwh"], "east_electricity_kwh": e["electricity_kwh"],
            "energy_intensity_kwh_m3": power / flow,
            "south_energy_intensity_kwh_m3": s["electricity_kwh"] / s["total_flow_m3"],
            "east_energy_intensity_kwh_m3": e["electricity_kwh"] / e["total_flow_m3"],
            "carbon_source_usage_kg": carbon, "carbon_intensity_kg_m3": carbon / flow,
            "flocculant_usage_kg": (s["flocculant_usage_kg"] or 0) + (e["flocculant_usage_kg"] or 0),
            "hypochlorite_usage_kg": hypo, "hypochlorite_intensity_kg_m3": hypo / flow,
            "tap_water_m3": (s["tap_water_m3"] or 0) + (e["tap_water_m3"] or 0),
            "sludge_t": (s["sludge_t"] or 0) + (e["sludge_t"] or 0),
            "pac_raw_kg": (s["pac_raw_kg"] or 0) + (e["pac_raw_kg"] or 0),
        }
        row_pair = f"{int(block['south_row'])},{int(block['east_row'])}"
        for metric_id, value in values_map.items():
            series_values[metric_id].append(value)
            sources[metric_id].append({"file": args.input, "sheet": sheet_name, "cell": row_pair,
                                       "formula": "derived from the two plant rows" if metric_id not in {"south_flow_m3", "east_flow_m3", "south_load_pct", "east_load_pct"} else None})

    summary_ws = values[adapter["summary_sheet"]]
    summary_cols = {key: column_index_from_string(value) for key, value in adapter["summary_columns"].items()}
    comparisons = []
    for cfg in profile.get("comparison_metrics", []):
        metric_id = cfg["metric_id"]
        col = summary_cols[metric_id]
        row25 = int(adapter["summary_rows"]["2025"]["total"])
        row26 = int(adapter["summary_rows"]["2026"]["total"])
        comparisons.append({"metric_id": metric_id, "label": cfg["label"], "unit": cfg["unit"],
                            "value_2025": numeric(summary_ws.cell(row25, col).value),
                            "value_2026": numeric(summary_ws.cell(row26, col).value),
                            "sources": [{"file": args.input, "sheet": adapter["summary_sheet"], "cell": f"{get_column_letter(col)}{row25}"},
                                        {"file": args.input, "sheet": adapter["summary_sheet"], "cell": f"{get_column_letter(col)}{row26}"}]})
    summary_2026 = {x["metric_id"]: x["value_2026"] for x in comparisons}

    observations, metrics = [], {}
    for metric_id, (unit, aggregation) in metric_defs.items():
        vals = series_values[metric_id]
        cached = summary_2026.get(metric_id)
        metric = {"metric_id": metric_id, "label": metric_id, "unit": unit, "aggregation": aggregation,
                  "periods": periods, "values": vals, "annual_cached_total": cached,
                  "annual_cached_average": None, "annual_total_formula": None, "annual_average_formula": None}
        metrics[metric_id] = metric
        for period, value, source in zip(periods, vals, sources[metric_id]):
            observations.append({"observation_id": f"OBS-{metric_id}-{period}", "metric_id": metric_id,
                                 "period": period, "value": value, "unit": unit, "quality_flag": "ok",
                                 "calculation_type": "derived_from_source_rows", "confirmation_status": "confirmed", "source": source})

    predefined = [
        {"severity":"warning","type":"summary_text_conflict","message":"汇总表文字说明中的处理量、同比方向及多项药剂总量与数值表不一致，正式计算以数值表为准。","source":{"sheet":adapter["summary_sheet"],"cell":"A15"}},
        {"severity":"warning","type":"pac_definition_conflict","metric_id":"pac_raw_kg","message":"月度PAC栏跨月混用液体PAC、固体PAC及混合/折算值，不能直接形成可比趋势。"},
        {"severity":"warning","type":"missing_cost_prices","message":"源文件未提供电价、药剂单价、水价、污泥处置单价或费用金额，本次不形成货币成本结论。"},
        {"severity":"warning","type":"missing_pump_selection_inputs","message":"工作簿目录提及加药泵选型，但未见选型表及峰值流量、扬程、运行小时等必要参数。"}
    ]
    write_json(args.output, {"schema_version":"0.1", "project":profile["project_name"], "analysis_year":year,
                             "sheet":sheet_name, "periods":periods, "metrics":metrics, "observations":observations,
                             "comparisons":comparisons, "predefined_quality_issues":predefined})
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--adapter", required=True)
    parser.add_argument("--profile", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    adapter = read_json(args.adapter)
    profile = read_json(args.profile)
    formulas = load_workbook(args.input, data_only=False, read_only=False)
    values = load_workbook(args.input, data_only=True, read_only=False)
    if adapter.get("layout") == "two_plant_monthly_blocks":
        normalize_two_plant(args, adapter, profile, formulas, values)
        return
    sheet_name = adapter["sheet"]
    if sheet_name not in formulas.sheetnames:
        raise SystemExit(f"Missing sheet: {sheet_name}")
    ws_f, ws_v = formulas[sheet_name], values[sheet_name]
    start_cell, end_cell = adapter["period_axis"].split(":")
    start_col = column_index_from_string("".join(filter(str.isalpha, start_cell)))
    end_col = column_index_from_string("".join(filter(str.isalpha, end_cell)))
    period_row = int("".join(filter(str.isdigit, start_cell)))
    year = int(profile["analysis_year"])
    periods = [month_label(year, ws_v.cell(period_row, col).value, index)
               for index, col in enumerate(range(start_col, end_col + 1), 1)]
    observations, metrics = [], {}
    annual_total_col = column_index_from_string(adapter["annual_total_column"])
    annual_avg_col = column_index_from_string(adapter["annual_average_column"])
    for metric_id, config in adapter["metrics"].items():
        row = int(config["row"])
        metric = {"metric_id": metric_id, **config, "periods": periods, "values": [],
                  "annual_cached_total": numeric(ws_v.cell(row, annual_total_col).value),
                  "annual_cached_average": numeric(ws_v.cell(row, annual_avg_col).value),
                  "annual_total_formula": ws_f.cell(row, annual_total_col).value,
                  "annual_average_formula": ws_f.cell(row, annual_avg_col).value}
        for period, col in zip(periods, range(start_col, end_col + 1)):
            raw = ws_v.cell(row, col).value
            value = numeric(raw)
            formula = ws_f.cell(row, col).value
            quality = "ok" if value is not None else "missing"
            observation_id = f"OBS-{metric_id}-{period}"
            item = {"observation_id": observation_id, "metric_id": metric_id, "period": period,
                    "value": value, "unit": config["unit"], "quality_flag": quality,
                    "calculation_type": "source_value" if not (isinstance(formula, str) and formula.startswith("=")) else "excel_formula_cache",
                    "confirmation_status": "confirmed" if value is not None else "to_confirm",
                    "source": {"file": args.input, "sheet": sheet_name,
                               "cell": f"{get_column_letter(col)}{row}", "formula": formula if formula != raw else None}}
            observations.append(item)
            metric["values"].append(value)
        metrics[metric_id] = metric
    write_json(args.output, {"schema_version": "0.1", "project": profile["project_name"],
                             "analysis_year": year, "sheet": sheet_name, "periods": periods,
                             "metrics": metrics, "observations": observations})


if __name__ == "__main__":
    main()
