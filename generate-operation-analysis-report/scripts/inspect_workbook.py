#!/usr/bin/env python3
import argparse
from openpyxl import load_workbook
from common import write_json


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input")
    parser.add_argument("output")
    args = parser.parse_args()
    wb = load_workbook(args.input, data_only=False, read_only=False)
    sheets = []
    for ws in wb.worksheets:
        formula_errors = []
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                if isinstance(cell.value, str) and cell.value in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"):
                    formula_errors.append({"cell": cell.coordinate, "value": cell.value})
        sheets.append({"name": ws.title, "rows": ws.max_row, "columns": ws.max_column,
                       "formula_count": formula_count, "formula_errors": formula_errors,
                       "charts": len(ws._charts), "images": len(ws._images)})
    write_json(args.output, {"schema_version": "0.1", "workbook": args.input, "sheets": sheets})


if __name__ == "__main__":
    main()
