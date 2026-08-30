#!/usr/bin/env python3
"""Export validation and drafting gaps as a reviewer-friendly workbook."""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def main():
    if len(sys.argv) != 4:
        raise SystemExit("Usage: export_pending_items.py validation.json drafted-content.json OUTPUT.xlsx")
    validation = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    draft = json.loads(Path(sys.argv[2]).read_text(encoding="utf-8"))
    rows = []
    for issue in validation.get("issues", []):
        source = issue.get("source", {})
        rows.append([issue.get("severity", "warning"), issue.get("type", ""),
                     issue.get("message") or issue.get("question", ""), source.get("sheet", ""),
                     source.get("answer_cell") or source.get("cell", ""), "待确认", ""])
    for chapter in draft.get("chapters", []):
        for item in chapter.get("to_confirm", []):
            if isinstance(item, str):
                text, source = item, {}
            else:
                text = item.get("item") or item.get("message") or item.get("question", "")
                source = item.get("source", {})
            rows.append(["warning", "draft_to_confirm", text, source.get("sheet", ""),
                         source.get("cell") or source.get("answer_cell", ""), "待确认", chapter.get("title", "")])
    unique, seen = [], set()
    for row in rows:
        key = tuple(str(value) for value in row[:5])
        if key not in seen:
            seen.add(key)
            unique.append(row)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "待确认事项"
    headers = ["级别", "类型", "事项", "来源工作表", "来源单元格", "状态", "影响章节"]
    sheet.append(headers)
    for row in unique:
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="24577F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [12, 24, 48, 28, 16, 14, 30]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    output = Path(sys.argv[3])
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


if __name__ == "__main__":
    main()
