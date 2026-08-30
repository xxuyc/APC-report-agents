#!/usr/bin/env python3
"""Export claim-to-source traceability as a reviewer-friendly workbook."""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: export_traceability.py traceability.json OUTPUT.xlsx")
    data = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "内容追溯"
    headers = ["结论ID", "章节", "类型", "报告表述", "来源文件", "来源工作表", "来源单元格"]
    sheet.append(headers)
    for claim in data.get("claims", []):
        sources = claim.get("sources", []) or [{}]
        for source in sources:
            sheet.append([claim.get("claim_id", ""), claim.get("chapter", ""), claim.get("kind", ""),
                          claim.get("text", ""), source.get("workbook", ""), source.get("sheet", ""),
                          source.get("cell") or source.get("answer_cell", "")])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="24577F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [16, 30, 14, 55, 30, 28, 16]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    output = Path(sys.argv[2])
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


if __name__ == "__main__":
    main()
