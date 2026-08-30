#!/usr/bin/env python3
"""Extract survey workbooks into the V0.2 auditable record contract."""

import json
import math
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


HEADER_ALIASES = {
    "question_id": ("题目编号", "问题编号", "编号"),
    "module": ("模块", "调研模块"),
    "submodule": ("子模块", "分类"),
    "question": ("设备/仪表名称", "设备/仪表类别", "设备仪表名称", "调研内容", "调研项", "调查内容", "问题"),
    "answer": ("调研结果", "填写内容", "答案", "现场情况", "调研情况", "问题描述"),
    "status_value": ("设备状态", "运行状态", "状态"),
    "required": ("是否必填", "必填", "重要程度", "调研等级"),
    "condition": ("适用条件", "条件", "填写条件", "备注"),
    "information_type": ("信息类型", "内容属性"),
    "confirmation_status": ("确认状态",),
    "evidence": ("证据来源及编号", "证据来源", "证据编号"),
    "verification_result": ("资料与现场是否一致", "资料核验结果"),
    "discrepancy": ("不一致情况说明", "差异说明"),
}

TARGET_OVERRIDES = {
    "05智能碳源投加调研": {"C-011", "C-012", "C-013", "C-014", "C-015", "C-016", "C-R017"},
    "06智能化学除磷调研": {"P-010", "P-011", "P-012", "P-013", "P-014", "P-R006"},
}

ROLE_MAP = {
    "现状事实": "current_fact", "现状": "current_fact", "背景资料": "background_fact",
    "现状问题": "problem", "问题": "problem", "建设需求": "target_requirement",
    "拟新增配置": "target_requirement", "待确认": "to_confirm",
}
STATUS_MAP = {"已确认": "confirmed", "确认": "confirmed", "待确认": "to_confirm", "不适用": "not_applicable"}


def normalized(value):
    return re.sub(r"\s+", "", str(value or ""))


def value(cell):
    if cell is None or cell.value is None:
        return ""
    if isinstance(cell.value, float) and math.isnan(cell.value):
        return ""
    return str(cell.value).strip()


def phase_normalize(text):
    result = str(text or "")
    replacements = (("1、2期", "一、二期"), ("1、2 期", "一、二期"), ("12期", "一、二期"),
                    ("12 期", "一、二期"), ("3期", "三期"), ("3 期", "三期"))
    for old, new in replacements:
        result = result.replace(old, new)
    return result


def map_header(row):
    labels = {normalized(cell.value): cell.column for cell in row if cell.value is not None}
    mapped = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            exact = next((column for label, column in labels.items() if normalized(alias) == label), None)
            partial = None if field == "status_value" else next(
                (column for label, column in labels.items() if normalized(alias) in label), None)
            if exact or partial:
                mapped[field] = exact or partial
                break
    if "question" in mapped and ("answer" in mapped or "status_value" in mapped):
        return mapped
    return {}


def find_headers(ws):
    return [(row[0].row, mapped) for row in ws.iter_rows(min_row=1, max_row=ws.max_row)
            if (mapped := map_header(row))]


def find_summary_header(ws):
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10)):
        labels = {normalized(cell.value): cell.column for cell in row if cell.value is not None}
        module = next((column for label, column in labels.items() if label == "模块"), None)
        applicable = next((column for label, column in labels.items() if "是否适用" in label), None)
        status = next((column for label, column in labels.items() if label == "状态"), None)
        if module and (applicable or status):
            return row[0].row, {"module": module, "applicable": applicable, "status": status}
    return None, {}


def infer_role(record):
    explicit = ROLE_MAP.get(record.get("information_type", ""))
    if explicit:
        return explicit
    sheet = record.get("source", {}).get("sheet", "")
    question_id = record.get("question_id", "")
    if question_id in TARGET_OVERRIDES.get(sheet, set()):
        return "target_requirement"
    combined = " ".join((record.get("submodule", ""), record.get("question", ""), question_id))
    if question_id.endswith("Q01") or any(token in combined for token in ("问题痛点", "主要问题", "问题描述")):
        return "problem"
    if any(token in combined for token in ("拟新增", "新增配置", "建设需求", "改造目标")):
        return "target_requirement"
    return "current_fact"


def infer_status(record):
    explicit = STATUS_MAP.get(record.get("confirmation_status", ""))
    if explicit:
        return explicit
    return "confirmed" if record.get("answer") else "to_confirm"


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: extract_survey.py INPUT.xlsx OUTPUT.json")
    source = Path(sys.argv[1])
    workbook = load_workbook(source, data_only=False)
    cached_workbook = load_workbook(source, data_only=True)
    sheets, records, formula_errors, module_status = [], [], [], []

    for ws in workbook.worksheets:
        headers = find_headers(ws)
        sheets.append({"name": ws.title, "max_row": ws.max_row, "max_column": ws.max_column,
                       "tables": [{"header_row": row, "columns": columns} for row, columns in headers]})
        summary_row, summary_columns = find_summary_header(ws)
        if summary_row:
            for row_number in range(summary_row + 1, ws.max_row + 1):
                module = value(ws.cell(row_number, summary_columns["module"]))
                if module:
                    module_status.append({
                        "module": module,
                        "applicable": value(ws.cell(row_number, summary_columns["applicable"])) if summary_columns["applicable"] else "",
                        "status": value(ws.cell(row_number, summary_columns["status"])) if summary_columns["status"] else "",
                        "source": {"sheet": ws.title, "row": row_number},
                    })
        cached_ws = cached_workbook[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                displayed, cached = value(cell), value(cached_ws[cell.coordinate])
                if "#REF!" in displayed or (cached.startswith("#") and "!" in cached):
                    formula_errors.append({"sheet": ws.title, "cell": cell.coordinate, "value": displayed})

        for header_index, (header_row, columns) in enumerate(headers):
            end_row = headers[header_index + 1][0] - 1 if header_index + 1 < len(headers) else ws.max_row
            for row_number in range(header_row + 1, end_row + 1):
                question_cell = ws.cell(row_number, columns["question"])
                question = value(question_cell)
                if not question:
                    continue
                answer_cell = ws.cell(row_number, columns.get("answer", columns.get("status_value")))
                components = {}
                for field in ("answer", "status_value"):
                    if columns.get(field):
                        item = value(ws.cell(row_number, columns[field]))
                        if item:
                            components[field] = item
                answer = components.get("answer", "")
                if components.get("status_value"):
                    answer = f"状态：{components['status_value']}" + (f"；{answer}" if answer else "")
                record = {
                    "source": {"workbook": source.name, "sheet": ws.title, "row": row_number,
                               "table_header_row": header_row, "question_cell": question_cell.coordinate,
                               "answer_cell": answer_cell.coordinate},
                    "question": question, "answer": answer, "answer_components": components,
                }
                for field, column in columns.items():
                    if field not in ("answer", "status_value"):
                        cell = ws.cell(row_number, column)
                        if field == "condition" and isinstance(cell.value, str) and cell.value.startswith("="):
                            record[field] = value(cached_ws[cell.coordinate])
                        else:
                            record[field] = value(cell)
                record["role"] = infer_role(record)
                record["confirmation"] = infer_status(record)
                record["normalized_answer"] = phase_normalize(answer)
                records.append(record)

    payload = {"schema_version": "0.2", "source_file": str(source), "sheets": sheets,
               "records": records, "module_status": module_status, "formula_errors": formula_errors}
    Path(sys.argv[2]).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
