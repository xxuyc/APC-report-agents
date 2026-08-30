#!/usr/bin/env python3
"""End-to-end smoke test for the deterministic parts of Harness V0.1."""

import copy
import json
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook
from docx import Document


SKILL_DIR = Path(__file__).resolve().parents[1]
PIPELINE = SKILL_DIR / "scripts" / "run_pipeline.py"
TEMPLATE = SKILL_DIR / "assets" / "survey-report-template.docx"


def run(*args):
    result = subprocess.run([sys.executable, str(PIPELINE), *map(str, args)],
                            capture_output=True, text=True, encoding="utf-8", errors="replace")
    if result.returncode:
        raise AssertionError(f"Command failed: {args}\n{result.stdout}\n{result.stderr}")
    return json.loads(result.stdout.strip().splitlines()[-1])


def run_expect_failure(*args):
    result = subprocess.run([sys.executable, str(PIPELINE), *map(str, args)],
                            capture_output=True, text=True, encoding="utf-8", errors="replace")
    assert result.returncode != 0
    return json.loads(result.stdout.strip().splitlines()[-1])


def create_workbook(path):
    workbook = Workbook()
    first = workbook.active
    first.title = "01项目基本信息"
    first.append(["题目编号", "调研内容", "调研结果", "备注", "是否必填"])
    first.append(["BASIC-001", "项目名称", "Harness 测试污水处理厂", "", "必填"])
    first.append(["BASIC-002", "设计规模", "5 万 m³/d", "", "必填"])
    control = workbook.create_sheet("03自控与数据")
    control.append(["题目编号", "模块", "子模块", "调研内容", "调研结果", "备注", "是否必填"])
    control.append(["CTRL-001", "自控系统", "控制方式", "现有控制方式", "主要依靠人工调节", "", "必填"])
    aeration = workbook.create_sheet("04精确曝气调研")
    aeration.append(["题目编号", "模块", "子模块", "调研内容", "调研结果", "备注", "是否必填"])
    aeration.append(["AIR-001", "精确曝气", "现状", "鼓风机数量", "5台", "", "必填"])
    carbon = workbook.create_sheet("05智能碳源投加调研")
    carbon.append(["题目编号", "模块", "子模块", "调研内容", "调研结果", "备注", "是否必填", "信息类型", "确认状态"])
    carbon.append(["C-011", "碳源", "设备配置", "拟新增加药泵数量", "3台", "", "必填", "建设需求", "已确认"])
    summary = workbook.create_sheet("调研完整度统计")
    summary.append(["模块", "是否适用", "状态"])
    summary.append(["项目基本信息", "是", "完成"])
    summary.append(["自控与数据", "是", "完成"])
    summary.append(["精确曝气", "否", "不适用"])
    summary.append(["智能碳源投加", "是", "完成"])
    workbook.save(path)


def create_supporting(path):
    document = Document()
    document.add_paragraph("项目名称：Harness 测试污水处理厂。设计规模为 5 万 m³/d。")
    document.save(path)


def main():
    with tempfile.TemporaryDirectory(prefix="survey-harness-") as temp:
        root = Path(temp)
        workbook = root / "survey.xlsx"
        supporting = root / "基本情况.docx"
        create_workbook(workbook)
        create_supporting(supporting)
        prepared = run("prepare", "--workspace", root / "workspace", "--project-id", "smoke-test",
                       "--project-name", "Harness 测试污水处理厂", "--input", workbook,
                       "--template", TEMPLATE, "--supporting", supporting)
        assert prepared["status"] == "ready_to_draft"
        run_dir = Path(prepared["run_dir"])
        manifest = json.loads((run_dir / "manifest.json").read_text(encoding="utf-8"))
        assert manifest["knowledge_mode"] == "disabled"
        assert manifest["knowledge_status"] == "disabled"
        model = json.loads((run_dir / "working" / "chapter-model.json").read_text(encoding="utf-8"))
        assert all(chapter.get("chapter_key") != "aeration" for chapter in model["chapters"])
        assert all(fact.get("role") != "target_requirement" for chapter in model["chapters"] for fact in chapter["facts"])
        source_index = json.loads((run_dir / "working" / "source-index.json").read_text(encoding="utf-8"))
        assert any(source["source_type"] == "supporting_document" for source in source_index["sources"])
        chapters = []
        for chapter in model["chapters"]:
            claims, paragraphs = [], []
            if chapter["facts"]:
                fact = chapter["facts"][0]
                claim_id = f"C-{chapter['chapter_id']}-001"
                claims.append({"claim_id": claim_id, "text": fact["value"], "kind": "fact",
                               "sources": [{"fact_id": fact["fact_id"]}]})
                paragraphs.append({"paragraph_id": f"P-{chapter['chapter_id']}-001",
                                   "text": f"{fact['value']}。",
                                   "claim_ids": [claim_id]})
            chapters.append({"chapter_id": chapter["chapter_id"], "title": chapter["title"],
                             "status": "completed" if paragraphs else "insufficient_facts",
                             "paragraphs": paragraphs, "claims": claims, "to_confirm": []})
        draft = {"schema_version": "0.2", "title": "Harness 测试现场调研现状评估", "chapters": chapters}
        invalid = copy.deepcopy(draft)
        next(claim for chapter in invalid["chapters"] for claim in chapter["claims"])["sources"] = []
        (run_dir / "working" / "drafted-content.json").write_text(
            json.dumps(invalid, ensure_ascii=False, indent=2), encoding="utf-8")
        rejected = run_expect_failure("finalize", "--run-dir", run_dir)
        assert rejected["status"] == "draft_failed"
        (run_dir / "working" / "drafted-content.json").write_text(
            json.dumps(draft, ensure_ascii=False, indent=2), encoding="utf-8")
        finalized = run("finalize", "--run-dir", run_dir)
        assert finalized["status"] == "completed_with_warnings"
        for name in ("调研报告_V0.2.docx", "待确认事项.xlsx", "内容追溯表.xlsx"):
            assert (run_dir / "output" / name).is_file()
        completed = run("record-qa", "--run-dir", run_dir, "--status", "passed", "--notes", "smoke test")
        assert completed["status"] in ("completed", "completed_with_warnings")

        # 用户批准的项目级金标准必须覆盖自动适用性判断，并允许把建设需求
        # 作为单独 requirements 输入；仍不得把它混入现状 facts。
        golden_dir = root / "workspace" / "golden-cases" / "golden-smoke"
        golden_dir.mkdir(parents=True)
        (golden_dir / "golden-profile.json").write_text(json.dumps({
            "schema_version": "0.1", "chapter_start": 2,
            "approved_chapter_keys": ["overview", "control", "aeration", "carbon", "summary"],
            "allow_confirmed_requirements": True,
            "module_applicability_overrides": {"精确曝气": True},
        }, ensure_ascii=False), encoding="utf-8")
        golden_prepared = run(
            "prepare", "--workspace", root / "workspace", "--project-id", "golden-smoke",
            "--project-name", "金标准测试厂", "--input", workbook, "--template", TEMPLATE,
        )
        golden_model = json.loads((Path(golden_prepared["run_dir"]) / "working" / "chapter-model.json")
                                  .read_text(encoding="utf-8"))
        assert golden_model["chapters"][0]["chapter_id"] == "3.2"
        assert any(chapter["chapter_key"] == "aeration" for chapter in golden_model["chapters"])
        carbon_chapter = next(chapter for chapter in golden_model["chapters"] if chapter["chapter_key"] == "carbon")
        assert carbon_chapter["requirements"]
        assert all(fact.get("role") != "target_requirement" for fact in carbon_chapter["facts"])
        print("Harness V0.2 smoke test: PASS")


if __name__ == "__main__":
    main()
